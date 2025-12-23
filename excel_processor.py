#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Application de traitement des fichiers CLIENT et BACKLOG
Permet de :
1. Confirmer les informations dans les colonnes T, U, V du fichier CLIENT
2. Générer des tableaux pour chaque demande de devancement
"""

import sys
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import traceback


class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Traitement CLIENT / BACKLOG")
        self.root.geometry("900x700")
        
        self.client_file = None
        self.backlog_file = None
        self.output_folder = None
        
        self.setup_ui()
    
    def setup_ui(self):
        """Créer l'interface graphique"""
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configuration du grid
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Titre
        title = ttk.Label(main_frame, text="Traitement CLIENT / BACKLOG", 
                         font=('Arial', 16, 'bold'))
        title.grid(row=0, column=0, columnspan=3, pady=10)
        
        # Section Fichiers
        files_frame = ttk.LabelFrame(main_frame, text="Fichiers", padding="10")
        files_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        files_frame.columnconfigure(1, weight=1)
        
        # Fichier CLIENT
        ttk.Label(files_frame, text="Fichier CLIENT:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.client_entry = ttk.Entry(files_frame, width=50)
        self.client_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Button(files_frame, text="Parcourir...", 
                  command=self.browse_client).grid(row=0, column=2, padx=5)
        
        # Fichier BACKLOG
        ttk.Label(files_frame, text="Fichier BACKLOG:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.backlog_entry = ttk.Entry(files_frame, width=50)
        self.backlog_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Button(files_frame, text="Parcourir...", 
                  command=self.browse_backlog).grid(row=1, column=2, padx=5)
        
        # Dossier de sortie
        ttk.Label(files_frame, text="Dossier de sortie:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.output_entry = ttk.Entry(files_frame, width=50)
        self.output_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Button(files_frame, text="Parcourir...", 
                  command=self.browse_output).grid(row=2, column=2, padx=5)
        
        # Section Options
        options_frame = ttk.LabelFrame(main_frame, text="Options de traitement", padding="10")
        options_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        self.confirm_info_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Confirmer les informations (colonnes T, U, V)", 
                       variable=self.confirm_info_var).grid(row=0, column=0, sticky=tk.W, pady=2)
        
        self.devancement_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Générer les tableaux de devancement", 
                       variable=self.devancement_var).grid(row=1, column=0, sticky=tk.W, pady=2)
        
        # Bouton de traitement
        process_btn = ttk.Button(main_frame, text="Lancer le traitement", 
                                command=self.process_files, style='Accent.TButton')
        process_btn.grid(row=3, column=0, columnspan=3, pady=20)
        
        # Zone de log
        log_frame = ttk.LabelFrame(main_frame, text="Résultats", padding="10")
        log_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        main_frame.rowconfigure(4, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, width=80)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # Barre de progression
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
    
    def log(self, message):
        """Ajouter un message au log"""
        self.log_text.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def browse_client(self):
        """Sélectionner le fichier CLIENT"""
        filename = filedialog.askopenfilename(
            title="Sélectionner le fichier CLIENT",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.client_file = filename
            self.client_entry.delete(0, tk.END)
            self.client_entry.insert(0, filename)
    
    def browse_backlog(self):
        """Sélectionner le fichier BACKLOG"""
        filename = filedialog.askopenfilename(
            title="Sélectionner le fichier BACKLOG",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.backlog_file = filename
            self.backlog_entry.delete(0, tk.END)
            self.backlog_entry.insert(0, filename)
    
    def browse_output(self):
        """Sélectionner le dossier de sortie"""
        folder = filedialog.askdirectory(title="Sélectionner le dossier de sortie")
        if folder:
            self.output_folder = folder
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, folder)
    
    def validate_inputs(self):
        """Valider les entrées"""
        if not self.client_file or not os.path.exists(self.client_file):
            messagebox.showerror("Erreur", "Veuillez sélectionner un fichier CLIENT valide")
            return False
        
        if not self.backlog_file or not os.path.exists(self.backlog_file):
            messagebox.showerror("Erreur", "Veuillez sélectionner un fichier BACKLOG valide")
            return False
        
        if not self.output_folder:
            # Utiliser le même dossier que le fichier CLIENT par défaut
            self.output_folder = os.path.dirname(self.client_file)
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, self.output_folder)
        
        if not os.path.exists(self.output_folder):
            try:
                os.makedirs(self.output_folder)
            except Exception as e:
                messagebox.showerror("Erreur", f"Impossible de créer le dossier: {str(e)}")
                return False
        
        return True
    
    def process_files(self):
        """Traiter les fichiers"""
        if not self.validate_inputs():
            return
        
        self.log_text.delete(1.0, tk.END)
        self.progress.start(10)
        
        try:
            self.log("Démarrage du traitement...")
            
            # Charger les fichiers
            self.log("Chargement du fichier CLIENT...")
            df_client = pd.read_excel(self.client_file)
            
            self.log("Chargement du fichier BACKLOG...")
            df_backlog = pd.read_excel(self.backlog_file)
            
            self.log(f"CLIENT: {len(df_client)} lignes chargées")
            self.log(f"BACKLOG: {len(df_backlog)} lignes chargées")
            
            # Traitement 1: Confirmer les informations
            if self.confirm_info_var.get():
                self.log("\n--- Confirmation des informations ---")
                self.confirm_information(df_client, df_backlog)
            
            # Traitement 2: Tableaux de devancement
            if self.devancement_var.get():
                self.log("\n--- Génération des tableaux de devancement ---")
                self.generate_devancement_tables(df_client, df_backlog)
            
            self.log("\n✓ Traitement terminé avec succès!")
            messagebox.showinfo("Succès", "Le traitement est terminé avec succès!")
            
        except Exception as e:
            error_msg = f"Erreur lors du traitement: {str(e)}\n{traceback.format_exc()}"
            self.log(f"\n✗ ERREUR: {error_msg}")
            messagebox.showerror("Erreur", f"Une erreur est survenue:\n{str(e)}")
        
        finally:
            self.progress.stop()
    
    def confirm_information(self, df_client, df_backlog):
        """Confirmer les informations dans les colonnes T, U, V"""
        self.log("Préparation des données pour confirmation...")
        
        # Créer un dictionnaire de correspondance à partir du BACKLOG
        # Clé: OrderNo, Valeurs: ConfirmedDate, Comment
        backlog_dict = {}
        for _, row in df_backlog.iterrows():
            order_no = str(row.get('OrderNo', '')).strip()
            if order_no and order_no != 'nan':
                if order_no not in backlog_dict:
                    backlog_dict[order_no] = {
                        'ConfirmedDate': row.get('ConfirmedDate'),
                        'Comment': row.get('Comment')
                    }
        
        self.log(f"Dictionnaire BACKLOG créé: {len(backlog_dict)} commandes")
        
        # Créer une copie du fichier CLIENT avec les informations confirmées
        wb = openpyxl.load_workbook(self.client_file)
        ws = wb.active
        
        # Trouver les index des colonnes
        col_numero_ar = None
        col_nouveau_delai = None
        col_nouveau_commentaire = None
        
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == 'Numero AR fournisseur':
                col_numero_ar = col_idx
            elif cell.value == 'Nouveau delai confirme par le fournisseur':
                col_nouveau_delai = col_idx
            elif cell.value == 'Nouveau Commentaire fournisseur':
                col_nouveau_commentaire = col_idx
        
        if not all([col_numero_ar, col_nouveau_delai, col_nouveau_commentaire]):
            self.log("✗ Colonnes T, U, V non trouvées dans le fichier CLIENT")
            return
        
        self.log(f"Colonnes identifiées: T={col_numero_ar}, U={col_nouveau_delai}, V={col_nouveau_commentaire}")
        
        # Parcourir les lignes et confirmer les informations
        confirmed_count = 0
        not_found_count = 0
        
        for row_idx in range(2, ws.max_row + 1):
            numero_ar = ws.cell(row_idx, col_numero_ar).value
            
            if numero_ar:
                numero_ar_str = str(numero_ar).strip()
                
                if numero_ar_str in backlog_dict:
                    # Confirmer les informations
                    backlog_info = backlog_dict[numero_ar_str]
                    
                    # Colonne U: Nouveau delai confirme
                    if pd.notna(backlog_info['ConfirmedDate']):
                        ws.cell(row_idx, col_nouveau_delai).value = backlog_info['ConfirmedDate']
                        ws.cell(row_idx, col_nouveau_delai).fill = PatternFill(
                            start_color='90EE90', end_color='90EE90', fill_type='solid'
                        )
                    
                    # Colonne V: Nouveau commentaire
                    if pd.notna(backlog_info['Comment']):
                        ws.cell(row_idx, col_nouveau_commentaire).value = backlog_info['Comment']
                        ws.cell(row_idx, col_nouveau_commentaire).fill = PatternFill(
                            start_color='90EE90', end_color='90EE90', fill_type='solid'
                        )
                    
                    confirmed_count += 1
                else:
                    not_found_count += 1
        
        # Sauvegarder le fichier
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(self.output_folder, f"CLIENT_confirme_{timestamp}.xlsx")
        wb.save(output_file)
        
        self.log(f"✓ Informations confirmées: {confirmed_count} lignes")
        self.log(f"  Non trouvées dans BACKLOG: {not_found_count} lignes")
        self.log(f"✓ Fichier sauvegardé: {output_file}")
    
    def generate_devancement_tables(self, df_client, df_backlog):
        """Générer des tableaux pour chaque demande de devancement"""
        self.log("Identification des demandes de devancement...")
        
        # Identifier les demandes de devancement
        # Une demande de devancement = ligne avec une date de livraison souhaitée antérieure à la date promise
        df_client['Date livraison souhaitee'] = pd.to_datetime(
            df_client['Date livraison souhaitee'], errors='coerce'
        )
        df_client['Date initiale promise'] = pd.to_datetime(
            df_client['Date initiale promise'], errors='coerce'
        )
        
        # Filtrer les devancement
        devancement_mask = (
            pd.notna(df_client['Date livraison souhaitee']) & 
            pd.notna(df_client['Date initiale promise']) &
            (df_client['Date livraison souhaitee'] < df_client['Date initiale promise'])
        )
        
        df_devancement = df_client[devancement_mask].copy()
        
        if len(df_devancement) == 0:
            self.log("✗ Aucune demande de devancement trouvée")
            return
        
        self.log(f"✓ {len(df_devancement)} demandes de devancement identifiées")
        
        # Créer un fichier Excel pour toutes les demandes
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(self.output_folder, f"Devancements_{timestamp}.xlsx")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Feuille récapitulative
            summary_data = []
            for idx, row in df_devancement.iterrows():
                summary_data.append({
                    'Symbole': row['Symbole'],
                    'Designation': row['Designation'],
                    'Numero AR': row.get('Numero AR fournisseur'),
                    'Date promise': row['Date initiale promise'],
                    'Date souhaitée': row['Date livraison souhaitee'],
                    'Jours de devancement': (row['Date initiale promise'] - row['Date livraison souhaitee']).days,
                    'Fournisseur': row.get('Fournisseur'),
                    'Quantite': row.get('Quantite CDA UM CDA')
                })
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Récapitulatif', index=False)
            
            # Formater la feuille récapitulative
            ws_summary = writer.sheets['Récapitulatif']
            self.format_worksheet(ws_summary)
            
            # Créer une feuille par demande (limitée aux 10 premières)
            max_sheets = min(10, len(df_devancement))
            for i, (idx, row) in enumerate(df_devancement.head(max_sheets).iterrows()):
                sheet_name = f"Dev_{i+1}_{str(row['Symbole'])[:10]}"
                
                # Rechercher les informations correspondantes dans BACKLOG
                numero_ar = str(row.get('Numero AR fournisseur', '')).strip()
                backlog_rows = df_backlog[df_backlog['OrderNo'].astype(str).str.strip() == numero_ar]
                
                # Créer le tableau de détail
                detail_data = {
                    'Information': [
                        'Symbole',
                        'Désignation',
                        'Numéro AR fournisseur',
                        'Date promise initiale',
                        'Date de livraison souhaitée',
                        'Jours de devancement demandé',
                        'Fournisseur',
                        'Quantité commandée',
                        '',
                        'Informations BACKLOG',
                        'Nombre de lignes trouvées',
                        'Quantité totale commandée',
                        'Quantité restante',
                        'Date de départ',
                        'Date confirmée',
                        'Commentaire'
                    ],
                    'Valeur': [
                        row['Symbole'],
                        row['Designation'],
                        row.get('Numero AR fournisseur'),
                        row['Date initiale promise'],
                        row['Date livraison souhaitee'],
                        (row['Date initiale promise'] - row['Date livraison souhaitee']).days,
                        row.get('Fournisseur'),
                        row.get('Quantite CDA UM CDA'),
                        '',
                        '',
                        len(backlog_rows),
                        backlog_rows['OrderedQuantity'].sum() if len(backlog_rows) > 0 else 0,
                        backlog_rows['RemainingQuantity'].sum() if len(backlog_rows) > 0 else 0,
                        backlog_rows['DepartureDate'].iloc[0] if len(backlog_rows) > 0 else '',
                        backlog_rows['ConfirmedDate'].iloc[0] if len(backlog_rows) > 0 else '',
                        backlog_rows['Comment'].iloc[0] if len(backlog_rows) > 0 else ''
                    ]
                }
                
                df_detail = pd.DataFrame(detail_data)
                df_detail.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Formater la feuille
                ws_detail = writer.sheets[sheet_name]
                self.format_worksheet(ws_detail)
        
        self.log(f"✓ Fichier de devancement créé: {output_file}")
        self.log(f"  - Feuille récapitulative: {len(df_devancement)} demandes")
        self.log(f"  - Feuilles de détail: {max_sheets} premières demandes")
    
    def format_worksheet(self, ws):
        """Formater une feuille Excel"""
        # Style pour les en-têtes
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formater la première ligne (en-têtes)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Appliquer les bordures à toutes les cellules
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border


def main():
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
